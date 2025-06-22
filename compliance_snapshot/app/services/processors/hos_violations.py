import pandas as pd
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook


def _process_df(df: pd.DataFrame, counter: Counter) -> None:
    """Normalise headers and update row-counts into the counter."""
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    if "violation_type" not in df.columns:
        raise ValueError(
            "Expected a 'Violation Type' column, got: "
            + ", ".join(df.columns)
        )
    counter.update(df["violation_type"].value_counts().to_dict())


def summarize(path: Path) -> dict:
    """
    Args:
        path: CSV or XLSX file where each row is one violation.
    Returns:
        {
          'total_violations': int,
          'violations_by_type': {str: int}
        }
    """
    counter = Counter()

    if path.suffix.lower() == ".csv":
        for chunk in pd.read_csv(path, chunksize=100_000):
            _process_df(chunk, counter)
    else:  # XLSX read-only streaming
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        try:
            idx = headers.index("Violation Type")
        except ValueError as e:
            raise ValueError("Expected 'Violation Type' header") from e
        for row in ws.iter_rows(min_row=2, values_only=True):
            counter[row[idx]] += 1

    total = int(sum(counter.values()))
    return {
        "total_violations": total,
        "violations_by_type": dict(counter)
    }
